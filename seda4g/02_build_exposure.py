#!/usr/bin/env python3
"""
02_build_exposure.py — build the county x wave high-speed-mobile exposure panel
from the NTIA SBDD / National Broadband Map archive.

NATIONAL mode (default): 50 states + DC x all 9 waves (Jun-2010 ... Jun-2014),
block-level build from the per-state Wireless CSVs (transtech==80 = terrestrial
mobile). This is the PRIMARY exposure measure: the NBM Analyze Tables leak fixed
wireless into their speed-tier columns (see validation_results.json) and are
used for validation only. Measures per county x wave (pop-share covered by >=1
provider, using max pct_blk_in_shape per block, weighted by census block pop):
    share_mobile_any     transtech==80
    share_mobile_t6      transtech==80 & maxaddown>=6  (>=3 Mbps adv. down)
    share_mobile_t7      transtech==80 & maxaddown>=7  (>=10 Mbps adv. down)
    share_mobile_t7_vzw  Verizon-only t7 (near-pure LTE)
    share_mobile_t6_vzw  Verizon-only t6 (grantees that coded LTE tier 6)
    share_anywless_t7    any wireless t7 incl. fixed (leakage diagnostics)
Pop weights: 2000 PL94-171 block pops for the two 2010 waves (2000-vintage
blocks), 2010 PL94-171 for 2011+ (vintage verified empirically: Jun-2011 DE
blocks match 2010 census 100%, 2000 census 44%).

Resumable: per-state-wave county aggregates cached in
data/external/seda4g/build/cov2_<ST>_<wave>.csv (+ diag2_ completeness file);
a crash loses at most one state-wave. After a successful cache write the
state-wave zip is DELETED (except the original 8-state pilot zips, which
03_validate.py re-reads) to stay under the ~20 GB disk budget.

Outputs (in seda4g/):
  exposure_county_panel_national.csv  county_fips x wave panel
  treatment_dates.csv                 first wave t6/t7 crosses 25/50/75%
  build_national_diagnostics.json     trajectory vs FCC, completeness, suspects

Usage:
  python3 02_build_exposure.py                 # all 9 waves + assemble
  python3 02_build_exposure.py --wave 2012-06  # one wave (no assemble)
  python3 02_build_exposure.py --assemble-only
  python3 02_build_exposure.py --pilot         # legacy pilot/analyze-table build
"""
import argparse
import glob
import io
import json
import os
import re
import sys
import traceback
import zipfile
from datetime import datetime, timezone

import numpy as np
import pandas as pd

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data", "external", "seda4g")
BUILD = os.path.join(DATA, "build")
OUT = os.path.join(ROOT, "seda4g")
os.makedirs(BUILD, exist_ok=True)

PILOT = ["DE", "CO", "MS", "OH", "WA", "TX", "MA", "MT"]
ST_FIPS = {"DE": "10", "CO": "08", "MS": "28", "OH": "39",
           "WA": "53", "TX": "48", "MA": "25", "MT": "30"}

STATES = ["AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "DC", "FL", "GA",
          "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA",
          "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY",
          "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC", "SD", "TN", "TX",
          "UT", "VT", "VA", "WA", "WV", "WI", "WY"]

# wave -> (outer zip name pattern, census block vintage of the file)
WAVES9 = {
    "2010-06": ("SBDD_{st}_Fall2010.zip", 2000),
    "2010-12": ("{st}-NBM-CSV-December-2010.zip", 2000),
    "2011-06": ("{st}-NBM-CSV-June-2011.zip", 2010),
    "2011-12": ("{st}-NBM-CSV-December-2011.zip", 2010),
    "2012-06": ("{st}-NBM-CSV-June-2012.zip", 2010),
    "2012-12": ("{st}-NBM-CSV-Dec-2012.zip", 2010),
    "2013-06": ("{st}-NBM-CSV-June-2013.zip", 2010),
    "2013-12": ("{st}-NBM-CSV-Dec-2013.zip", 2010),
    "2014-06": ("{st}-NBM-CSV-June-2014.zip", 2010),
}
# alternate outer zip names (01_download.sh fallback scheme)
WAVES9_ALT = {
    "2010-06": "{st}-NBM-CSV-June-2010.zip",
    "2010-12": "{st}-NBM-CSV-Dec-2010.zip",
    "2011-06": "{st}-NBM-CSV-Jun-2011.zip",
    "2011-12": "{st}-NBM-CSV-Dec-2011.zip",
    "2012-06": "{st}-NBM-CSV-Jun-2012.zip",
    "2012-12": "{st}-NBM-CSV-December-2012.zip",
    "2013-06": "{st}-NBM-CSV-Jun-2013.zip",
    "2013-12": "{st}-NBM-CSV-December-2013.zip",
    "2014-06": "{st}-NBM-CSV-Jun-2014.zip",
}

# zips 03_validate.py / the pilot build still read directly — never delete
PRESERVE_ZIPS = set()
for _st in PILOT:
    PRESERVE_ZIPS.add(f"SBDD_{_st}_Fall2010.zip")
    PRESERVE_ZIPS.add(f"{_st}-NBM-CSV-December-2010.zip")
    PRESERVE_ZIPS.add(f"{_st}-NBM-CSV-June-2014.zip")
for _st in ["MT", "MS", "CO", "MA"]:
    PRESERVE_ZIPS.add(f"{_st}-NBM-CSV-Dec-2012.zip")

VZW = re.compile(r"verizon|cellco", re.I)  # Cellco Partnership dba Verizon Wireless

# ----------------------------------------------------------------------------
# 1. Census block populations
# ----------------------------------------------------------------------------

def blockpop_2010(st):
    """2010 PL94-171 geo file: SUMLEV 750 rows carry POP100. Fixed-width."""
    cache = os.path.join(BUILD, f"blockpop2010_{st.lower()}.csv.gz")
    if os.path.exists(cache):
        d = pd.read_csv(cache, dtype={"block": str})
        return dict(zip(d["block"], d["pop"]))
    zf = zipfile.ZipFile(os.path.join(DATA, "census", f"{st.lower()}2010.pl.zip"))
    geoname = [n for n in zf.namelist() if "geo2010" in n][0]
    out = {}
    with zf.open(geoname) as f:
        for raw in io.TextIOWrapper(f, encoding="latin-1"):
            if raw[8:11] != "750":
                continue
            block = raw[27:32] + raw[54:60] + raw[61:65]  # state+county+tract+block
            out[block] = int(raw[318:327])
    tmp = cache + ".tmp.gz"
    pd.DataFrame({"block": list(out), "pop": list(out.values())}).to_csv(
        tmp, index=False, compression="gzip")
    os.replace(tmp, cache)
    return out


def blockpop_2000(st):
    """2000 PL94-171: geo file (fixed width) gives LOGRECNO->block GEOID at
    SUMLEV 750; part-1 file (comma) gives P0010001 (5th and 6th fields)."""
    cache = os.path.join(BUILD, f"blockpop2000_{st.lower()}.csv.gz")
    if os.path.exists(cache):
        d = pd.read_csv(cache, dtype={"block": str})
        return dict(zip(d["block"], d["pop"]))
    geo_zf = zipfile.ZipFile(os.path.join(DATA, "census", f"{st.lower()}geo.upl.zip"))
    p1_zf = zipfile.ZipFile(os.path.join(DATA, "census", f"{st.lower()}00001.upl.zip"))
    logrec2block = {}
    with geo_zf.open(geo_zf.namelist()[0]) as f:
        for raw in io.TextIOWrapper(f, encoding="latin-1"):
            if raw[8:11] != "750":
                continue
            logrec = raw[18:25]
            block = raw[29:34] + raw[55:61] + raw[62:66]
            logrec2block[logrec] = block
    out = {}
    with p1_zf.open(p1_zf.namelist()[0]) as f:
        for raw in io.TextIOWrapper(f, encoding="latin-1"):
            parts = raw.split(",")
            logrec = parts[4].zfill(7)
            if logrec in logrec2block:
                out[logrec2block[logrec]] = int(parts[5])
    tmp = cache + ".tmp.gz"
    pd.DataFrame({"block": list(out), "pop": list(out.values())}).to_csv(
        tmp, index=False, compression="gzip")
    os.replace(tmp, cache)
    return out

# ----------------------------------------------------------------------------
# 2. zip member extraction (Deflate64-aware)
# ----------------------------------------------------------------------------

def _extract_member(zf, info, dest):
    """Extract a zip member to path `dest`, with a manual Deflate64
    (compress_type 9) fallback — several SBDD zips (e.g. TX Jun-2014) use
    Deflate64, unsupported by the stdlib (and by macOS unzip/bsdtar)."""
    if info.compress_type != 9:
        with zf.open(info) as src, open(dest, "wb") as out:
            while True:
                chunk = src.read(1 << 24)
                if not chunk:
                    break
                out.write(chunk)
        return
    import inflate64
    fp = zf.fp
    fp.seek(info.header_offset)
    hdr = fp.read(30)
    n_name = int.from_bytes(hdr[26:28], "little")
    n_extra = int.from_bytes(hdr[28:30], "little")
    fp.seek(info.header_offset + 30 + n_name + n_extra)
    inflater = inflate64.Inflater()
    remaining = info.compress_size
    with open(dest, "wb") as out:
        while remaining > 0:
            chunk = fp.read(min(1 << 24, remaining))
            remaining -= len(chunk)
            out.write(inflater.inflate(chunk))

# ----------------------------------------------------------------------------
# 3. SBDD wireless CSV -> block coverage (one state x wave) — national version
# ----------------------------------------------------------------------------

MEASURES = {
    "mobile_any":    lambda d: d["transtech"] == 80,
    "mobile_t6":     lambda d: (d["transtech"] == 80) & (d["maxaddown"] >= 6),
    "mobile_t7":     lambda d: (d["transtech"] == 80) & (d["maxaddown"] >= 7),
    "mobile_t7_vzw": lambda d: ((d["transtech"] == 80) & (d["maxaddown"] >= 7)
                                & d["provider"].str.contains(VZW)),
    "mobile_t6_vzw": lambda d: ((d["transtech"] == 80) & (d["maxaddown"] >= 6)
                                & d["provider"].str.contains(VZW)),
    "anywless_t7":   lambda d: d["maxaddown"] >= 7,   # incl. fixed wireless
}

BLOCKCOL_CANDIDATES = ["fullfipsid", "censusblock_fips", "censusblock",
                       "blockid", "block_fips", "block_id"]


def _outer_zip_path(st, wave):
    for pat in (WAVES9[wave][0], WAVES9_ALT[wave]):
        p = os.path.join(DATA, "states", pat.format(st=st))
        if os.path.exists(p):
            return p
    return None


def read_wireless_national(st, wave):
    """Stream one state x wave Wireless CSV. Returns
    (dict measure -> Series(block -> max pct_blk_in_shape), diag dict updates,
     set of all block ids seen)."""
    outer_path = _outer_zip_path(st, wave)
    if outer_path is None:
        raise FileNotFoundError(f"no zip for {st} {wave}")
    outer = zipfile.ZipFile(outer_path)
    cand = [i for i in outer.infolist()
            if re.search("wireless", i.filename, re.I)
            and i.filename.lower().endswith(".zip")]
    tag = f"{st}_{wave}"
    tmp_csv = os.path.join(BUILD, f"_tmp_{tag}.csv")
    if cand:
        info = max(cand, key=lambda i: i.file_size)  # largest wireless zip
        tmp_inner = os.path.join(BUILD, f"_tmp_{tag}_inner.zip")
        _extract_member(outer, info, tmp_inner)
        inner = zipfile.ZipFile(tmp_inner)
        datanames = [n for n in inner.namelist()
                     if n.lower().endswith((".csv", ".txt"))]
        if not datanames:
            inner.close()
            os.remove(tmp_inner)
            raise ValueError(f"wireless zip has no csv/txt for {st} {wave}")
        _extract_member(inner, inner.getinfo(datanames[0]), tmp_csv)
        inner.close()
        os.remove(tmp_inner)
    else:
        flat = [i for i in outer.infolist()
                if re.search("wireless", i.filename, re.I)
                and i.filename.lower().endswith((".csv", ".txt"))]
        if not flat:
            raise ValueError(f"no wireless member in {os.path.basename(outer_path)}")
        _extract_member(outer, max(flat, key=lambda i: i.file_size), tmp_csv)
    outer.close()

    with open(tmp_csv, encoding="latin-1") as f:
        head = f.readline()
    sep = "|" if head.count("|") > head.count(",") else ","
    hdr = [c.strip().lower().strip('"') for c in head.strip().split(sep)]
    blockcol = next((c for c in BLOCKCOL_CANDIDATES if c in hdr), None)
    if blockcol is None:
        blockcol = next((c for c in hdr if "fips" in c), None)
    if blockcol is None or "transtech" not in hdr or "maxaddown" not in hdr:
        os.remove(tmp_csv)
        raise ValueError(f"unrecognized schema {st} {wave}: {hdr[:14]}")
    has_pct = "pct_blk_in_shape" in hdr
    namecols = [c for c in ("provname", "dbaname", "hoconame", "providername")
                if c in hdr]
    use = [blockcol, "transtech", "maxaddown"] + namecols
    if has_pct:
        use.append("pct_blk_in_shape")

    partials = {m: [] for m in MEASURES}
    allblocks = set()
    n_rows = 0
    rawmax = 0.0
    for ch in pd.read_csv(tmp_csv, sep=sep, dtype=str, encoding="latin-1",
                          on_bad_lines="skip", header=0, names=hdr,
                          usecols=use, chunksize=1_500_000):
        n_rows += len(ch)
        ch = ch.rename(columns={blockcol: "block"})
        ch["block"] = ch["block"].str.replace('"', "").str.strip()
        # numeric-stored FIPS drop the leading zero in some grantee files
        ch.loc[ch["block"].str.len() == 14, "block"] = \
            "0" + ch.loc[ch["block"].str.len() == 14, "block"]
        ch = ch[ch["block"].str.len() == 15]
        ch["transtech"] = pd.to_numeric(ch["transtech"], errors="coerce")
        ch["maxaddown"] = pd.to_numeric(ch["maxaddown"], errors="coerce")
        if has_pct:
            ch["pct"] = pd.to_numeric(ch["pct_blk_in_shape"],
                                      errors="coerce").fillna(0.0)
        else:
            ch["pct"] = 1.0
        if len(ch):
            rawmax = max(rawmax, float(ch["pct"].max()))
        prov = ch[namecols[0]].fillna("") if namecols else pd.Series("", index=ch.index)
        for c in namecols[1:]:
            prov = prov + "|" + ch[c].fillna("")
        ch["provider"] = prov
        allblocks.update(ch["block"].unique())
        for m, mask in MEASURES.items():
            sub = ch[mask(ch)]
            if len(sub):
                partials[m].append(sub.groupby("block")["pct"].max())
    os.remove(tmp_csv)
    out = {}
    # Some waves store percent (0-100), others fraction (0-1). Decide once per
    # file from the largest pct seen in any row.
    for m, parts in partials.items():
        if parts:
            s = pd.concat(parts).groupby(level=0).max()
        else:
            s = pd.Series(dtype=float)
        if rawmax > 1.5 and len(s):
            s = s / 100.0
        out[m] = s.clip(0, 1)
    diag = dict(csv_rows=n_rows, pct_units="percent" if rawmax > 1.5 else "fraction",
                has_pct_col=has_pct, namecols="|".join(namecols),
                blockcol=blockcol)
    return out, diag, allblocks


def county_coverage_national(block_pct, pops):
    """block_pct: dict measure -> Series(block -> max pct_blk_in_shape).
    County pop-weighted coverage shares (max pct per block, '_wt'), any-overlap
    ('_any'), plus block counts."""
    blocks = pd.DataFrame({"block": list(pops), "pop": list(pops.values())})
    blocks["county"] = blocks["block"].str[:5]
    cty = blocks.groupby("county")["pop"].sum().rename("county_pop")
    res = pd.DataFrame(index=cty.index)
    res["county_pop"] = cty
    res["n_blocks_total"] = blocks.groupby("county")["block"].count()
    for name, bm in block_pct.items():
        m = blocks.merge(bm.rename("pct"), left_on="block",
                         right_index=True, how="left")
        m["pct"] = m["pct"].fillna(0.0)
        denom = cty.replace(0, np.nan)
        res[name + "_wt"] = (m["pop"] * m["pct"]).groupby(m["county"]).sum() / denom
        res[name + "_any"] = (m["pop"] * (m["pct"] > 0)).groupby(m["county"]).sum() / denom
        if name == "mobile_any":
            res["n_blocks_covered"] = (m["pct"] > 0).groupby(m["county"]).sum()
    return res.reset_index()


def process_state_wave(args):
    """Worker: build + cache county aggregates for one state x wave.
    Returns a status dict. Deletes the state-wave zip on success (unless
    preserved or keep_zips)."""
    st, wave, keep_zips = args
    ccache = os.path.join(BUILD, f"cov2_{st}_{wave}.csv")
    dcache = os.path.join(BUILD, f"diag2_{st}_{wave}.csv")
    fmark = os.path.join(BUILD, f"fail2_{st}_{wave}.json")
    if os.path.exists(ccache) and os.path.exists(dcache):
        return dict(state=st, wave=wave, status="cached")
    vintage = WAVES9[wave][1]
    outer_path = _outer_zip_path(st, wave)
    if outer_path is None:
        return dict(state=st, wave=wave, status="missing_zip")
    try:
        pops = blockpop_2000(st) if vintage == 2000 else blockpop_2010(st)
        block_pct, fdiag, allblocks = read_wireless_national(st, wave)
        cov = county_coverage_national(block_pct, pops)
        cov.insert(0, "wave", wave)
        cov.insert(1, "state", st)
        tot = sum(pops.values())
        pop_incsv = sum(p for b, p in pops.items() if b in allblocks)
        d = dict(wave=wave, state=st, vintage=vintage, state_pop=tot,
                 n_blocks_census=len(pops), n_blocks_in_csv=len(allblocks),
                 pop_share_in_csv=(pop_incsv / tot) if tot else np.nan,
                 csv_blocks_not_in_census=len(allblocks - set(pops)),
                 **fdiag)
        pd.DataFrame([d]).to_csv(dcache + ".tmp", index=False)
        cov.to_csv(ccache + ".tmp", index=False)
        os.replace(dcache + ".tmp", dcache)
        os.replace(ccache + ".tmp", ccache)
        if os.path.exists(fmark):
            os.remove(fmark)
        if not keep_zips and os.path.basename(outer_path) not in PRESERVE_ZIPS:
            os.remove(outer_path)
        return dict(state=st, wave=wave, status="built", rows=fdiag["csv_rows"])
    except Exception as e:
        err = dict(state=st, wave=wave, status="parse_error",
                   error=f"{type(e).__name__}: {e}",
                   traceback=traceback.format_exc()[-2000:])
        with open(fmark, "w") as f:
            json.dump(err, f, indent=2)
        return err


def build_wave(wave, jobs=3, keep_zips=False):
    args = [(st, wave, keep_zips) for st in STATES]
    results = []
    if jobs > 1:
        from multiprocessing import Pool
        with Pool(jobs, maxtasksperchild=4) as pool:
            for r in pool.imap_unordered(process_state_wave, args):
                results.append(r)
                if r["status"] != "cached":
                    print(f"  {r['state']} {wave}: {r['status']}"
                          + (f" ({r.get('error','')})" if "error" in r else ""),
                          flush=True)
    else:
        for a in args:
            r = process_state_wave(a)
            results.append(r)
            if r["status"] != "cached":
                print(f"  {r['state']} {wave}: {r['status']}", flush=True)
    return results

# ----------------------------------------------------------------------------
# 4. Assemble national outputs
# ----------------------------------------------------------------------------

FCC_BENCHMARKS = {
    "2010-06_any_lte": 0.0,
    "2012-10_any_lte_mosaik": 0.856,
    "2014-01_any_lte": 0.985,
    "note": ("FCC Mobile Competition Reports 16/17/19: any-LTE pop coverage. "
             "share_mobile_t7 is a >=10Mbps-advertised proxy and should track "
             "these from below (and sit below the validation exercise's "
             "any-WIRELESS t7 numbers, which include fixed-wireless WISPs)."),
}

SHARE_RENAME = {
    "county": "county_fips",
    "mobile_any_wt": "share_mobile_any",
    "mobile_t6_wt": "share_mobile_t6",
    "mobile_t7_wt": "share_mobile_t7",
    "mobile_t7_vzw_wt": "share_mobile_t7_vzw",
    "mobile_t6_vzw_wt": "share_mobile_t6_vzw",
    "anywless_t7_wt": "share_anywless_t7",
}
SHARES = ["share_mobile_any", "share_mobile_t6", "share_mobile_t7",
          "share_mobile_t7_vzw", "share_mobile_t6_vzw", "share_anywless_t7"]


def assemble():
    waves = list(WAVES9)
    frames, diags, failed = [], [], []
    for wave in waves:
        for st in STATES:
            ccache = os.path.join(BUILD, f"cov2_{st}_{wave}.csv")
            dcache = os.path.join(BUILD, f"diag2_{st}_{wave}.csv")
            fmark = os.path.join(BUILD, f"fail2_{st}_{wave}.json")
            if os.path.exists(ccache) and os.path.exists(dcache):
                frames.append(pd.read_csv(ccache, dtype={"county": str}))
                diags.append(pd.read_csv(dcache).iloc[0].to_dict())
            elif os.path.exists(fmark):
                with open(fmark) as f:
                    e = json.load(f)
                failed.append(dict(state=st, wave=wave, reason="parse_error",
                                   error=e.get("error", "")))
            else:
                failed.append(dict(state=st, wave=wave,
                                   reason="download_failed_or_unprocessed"))
    panel = pd.concat(frames, ignore_index=True)
    panel = panel.rename(columns=SHARE_RENAME)
    cols = (["wave", "state", "county_fips", "county_pop",
             "n_blocks_total", "n_blocks_covered"] + SHARES)
    panel = panel[cols].sort_values(["wave", "county_fips"])
    panel.to_csv(os.path.join(OUT, "exposure_county_panel_national.csv"),
                 index=False, float_format="%.6f")
    print(f"wrote exposure_county_panel_national.csv "
          f"({len(panel)} rows, {panel['county_fips'].nunique()} counties, "
          f"{panel['wave'].nunique()} waves)")

    # ---- treatment dates -------------------------------------------------
    wnum = {w: i for i, w in enumerate(waves)}
    p = panel.copy()
    p["wnum"] = p["wave"].map(wnum)
    base = (p.sort_values("wnum").groupby("county_fips")
            .agg(county_pop=("county_pop", "last"), state=("state", "last"),
                 n_waves_observed=("wave", "nunique")))
    td = base.copy()
    for meas, col in [("t6", "share_mobile_t6"), ("t7", "share_mobile_t7")]:
        for thr in (0.25, 0.50, 0.75):
            lab = f"{meas}_{int(thr*100)}"
            first = (p[p[col] >= thr].groupby("county_fips")["wnum"].min()
                     .map(lambda i: waves[int(i)]))
            td[f"first_{lab}"] = first
            td[f"never_{lab}_by_jun2014"] = first.reindex(td.index).isna()
    td = td.reset_index()
    td.to_csv(os.path.join(OUT, "treatment_dates.csv"), index=False)
    print(f"wrote treatment_dates.csv ({len(td)} counties)")

    # ---- diagnostics ------------------------------------------------------
    D = {"built_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
         "n_state_waves_expected": len(STATES) * len(waves),
         "n_state_waves_built": len(diags),
         "failed_state_waves": failed,
         "fcc_benchmarks": FCC_BENCHMARKS}

    # (i) national pop-weighted trajectory
    traj = []
    for w, g in panel.groupby("wave"):
        g = g[g["county_pop"] > 0]
        row = {"wave": w, "n_counties": int(len(g)),
               "pop_covered_m": round(g["county_pop"].sum() / 1e6, 2),
               "states_missing": sorted(set(STATES) - set(g["state"]))}
        for s in SHARES:
            row[s] = round(float(np.average(g[s].fillna(0),
                                            weights=g["county_pop"])), 4)
        traj.append(row)
    D["national_trajectory_popwt"] = traj

    # (ii) per-state completeness (pop share of state in wireless CSV rows)
    dd = pd.DataFrame(diags)
    comp = dd.pivot_table(index="state", columns="wave",
                          values="pop_share_in_csv").round(4)
    D["per_state_pop_share_in_csv"] = {
        st: {w: (None if pd.isna(v) else float(v)) for w, v in r.items()}
        for st, r in comp.iterrows()}
    low = dd[dd["pop_share_in_csv"] < 0.90][
        ["state", "wave", "pop_share_in_csv"]].round(4)
    D["state_waves_pop_share_below_90pct"] = low.to_dict("records")

    # (iii) tier-coding suspects: statewide t7 < 5% while t6 > 30%
    sw = (panel[panel["county_pop"] > 0].groupby(["state", "wave"])
          .apply(lambda g: pd.Series({
              "t7": np.average(g["share_mobile_t7"].fillna(0), weights=g["county_pop"]),
              "t6": np.average(g["share_mobile_t6"].fillna(0), weights=g["county_pop"]),
              "any": np.average(g["share_mobile_any"].fillna(0), weights=g["county_pop"]),
          }), include_groups=False).reset_index())
    susp = sw[(sw["t7"] < 0.05) & (sw["t6"] > 0.30)].round(4)
    D["tier_coding_suspects_t7lt5_t6gt30"] = susp.to_dict("records")
    sw.round(4).to_csv(os.path.join(BUILD, "statewide_shares.csv"), index=False)

    # (iv) treatment cohort distribution (first_t7_50)
    coh = td["first_t7_50"].fillna("never_by_jun2014").value_counts().sort_index()
    D["cohort_counts_first_t7_50"] = {k: int(v) for k, v in coh.items()}
    cohp = (td.dropna(subset=["county_pop"])
            .assign(c=td["first_t7_50"].fillna("never_by_jun2014"))
            .groupby("c")["county_pop"].sum())
    D["cohort_pop_share_first_t7_50"] = {
        k: round(float(v / cohp.sum()), 4) for k, v in cohp.sort_index().items()}

    with open(os.path.join(OUT, "build_national_diagnostics.json"), "w") as f:
        json.dump(D, f, indent=2, default=str)
    print("wrote build_national_diagnostics.json")
    return D

# ----------------------------------------------------------------------------
# 5. LEGACY pilot build (kept for 03_validate.py; see git history for context)
# ----------------------------------------------------------------------------

WAVE_FILES = {
    "2010-06": ("states/SBDD_{st}_Fall2010.zip", r"Wireless.*fall2010", 2000),
    "2010-12": ("states/{st}-NBM-CSV-December-2010.zip",
                r"Wireless-CSV-December-2010", 2000),
    "2012-12": ("states/{st}-NBM-CSV-Dec-2012.zip",
                r"WIRELESS-CSV-Dec-2012", 2010),
    "2014-06": ("states/{st}-NBM-CSV-June-2014.zip",
                r"Wireless-CSV-JUN-2014", 2010),
}

LEGACY_MEASURES = {k: v for k, v in MEASURES.items() if k != "mobile_t6_vzw"}


def read_wireless(st, wave):
    """Legacy pilot reader (4 waves x 8 states)."""
    outer_pat, inner_re, _ = WAVE_FILES[wave]
    outer = zipfile.ZipFile(os.path.join(DATA, outer_pat.format(st=st)))
    inner_names = [n for n in outer.namelist()
                   if re.search(inner_re, n, re.I) and n.lower().endswith(".zip")]
    if not inner_names:
        raise FileNotFoundError(f"no wireless zip for {st} {wave}")
    tmp_inner = os.path.join(BUILD, "_tmp_inner.zip")
    tmp_csv = os.path.join(BUILD, "_tmp_wireless.csv")
    _extract_member(outer, outer.getinfo(inner_names[0]), tmp_inner)
    inner = zipfile.ZipFile(tmp_inner)
    dataname = [n for n in inner.namelist()
                if n.lower().endswith((".csv", ".txt"))][0]
    _extract_member(inner, inner.getinfo(dataname), tmp_csv)
    inner.close()

    with open(tmp_csv, encoding="latin-1") as f:
        head = f.readline()
    sep = "|" if head.count("|") > head.count(",") else ","
    hdr = [c.strip().lower().strip('"') for c in head.strip().split(sep)]
    blockcol = "fullfipsid" if "fullfipsid" in hdr else "censusblock_fips"
    namecols = [c for c in ("provname", "dbaname", "hoconame") if c in hdr]
    use = [blockcol, "pct_blk_in_shape", "transtech", "maxaddown"] + namecols

    partials = {m: [] for m in LEGACY_MEASURES}
    allblocks = set()
    n_rows = 0
    for ch in pd.read_csv(tmp_csv, sep=sep, dtype=str, encoding="latin-1",
                          on_bad_lines="skip", header=0, names=hdr,
                          usecols=use, chunksize=2_000_000):
        n_rows += len(ch)
        ch = ch.rename(columns={blockcol: "block"})
        ch["block"] = ch["block"].str.replace('"', "").str.strip()
        ch = ch[ch["block"].str.len() == 15]
        ch["transtech"] = pd.to_numeric(ch["transtech"], errors="coerce")
        ch["maxaddown"] = pd.to_numeric(ch["maxaddown"], errors="coerce")
        ch["pct"] = pd.to_numeric(ch["pct_blk_in_shape"],
                                  errors="coerce").fillna(0.0)
        prov = ch[namecols[0]].fillna("") if namecols else ""
        for c in namecols[1:]:
            prov = prov + "|" + ch[c].fillna("")
        ch["provider"] = prov
        allblocks.update(ch["block"].unique())
        for m, mask in LEGACY_MEASURES.items():
            sub = ch[mask(ch)]
            if len(sub):
                partials[m].append(sub.groupby("block")["pct"].max())
    os.remove(tmp_inner)
    os.remove(tmp_csv)
    out = {}
    gmax = max((p.max() for parts in partials.values() for p in parts
                if len(p)), default=0.0)
    for m, parts in partials.items():
        if parts:
            s = pd.concat(parts).groupby(level=0).max()
        else:
            s = pd.Series(dtype=float)
        if gmax > 1.5 and len(s):
            s = s / 100.0
        out[m] = s.clip(0, 1)
    return out, n_rows, allblocks


def county_coverage(block_pct, pops):
    blocks = pd.DataFrame({"block": list(pops), "pop": list(pops.values())})
    blocks["county"] = blocks["block"].str[:5]
    cty = blocks.groupby("county")["pop"].sum().rename("county_pop")
    res = pd.DataFrame(index=cty.index)
    res["county_pop"] = cty
    for name, bm in block_pct.items():
        m = blocks.merge(bm.rename("pct"), left_on="block",
                         right_index=True, how="left")
        m["pct"] = m["pct"].fillna(0.0)
        res[name + "_wt"] = (m["pop"] * m["pct"]).groupby(m["county"]).sum() / cty
        res[name + "_any"] = (m["pop"] * (m["pct"] > 0)).groupby(m["county"]).sum() / cty
    return res.reset_index()


def build_block_panel():
    rows = []
    blockstats = []
    for wave, (_, _, vintage) in WAVE_FILES.items():
        for st in PILOT:
            ccache = os.path.join(BUILD, f"cov_{st}_{wave}.csv")
            dcache = os.path.join(BUILD, f"diag_{st}_{wave}.csv")
            if os.path.exists(ccache) and os.path.exists(dcache):
                rows.append(pd.read_csv(ccache, dtype={"county": str}))
                blockstats.append(pd.read_csv(dcache).iloc[0].to_dict())
                continue
            print(f"  block build {st} {wave}", flush=True)
            pops = blockpop_2000(st) if vintage == 2000 else blockpop_2010(st)
            try:
                block_pct, n_rows, allblocks = read_wireless(st, wave)
            except FileNotFoundError as e:
                print("   MISSING:", e)
                continue
            cov = county_coverage(block_pct, pops)
            cov.insert(0, "wave", wave)
            cov.insert(1, "state", st)
            cov.to_csv(ccache, index=False)
            rows.append(cov)
            tot = sum(pops.values())
            pop_incsv = sum(p for b, p in pops.items() if b in allblocks)
            n_unmatched = len(allblocks - set(pops))
            d = dict(
                wave=wave, state=st, vintage=vintage,
                state_pop=tot, n_blocks_census=len(pops),
                n_blocks_in_csv=len(allblocks),
                pop_share_in_csv=pop_incsv / tot,
                csv_blocks_not_in_census=n_unmatched,
                csv_rows=n_rows)
            pd.DataFrame([d]).to_csv(dcache, index=False)
            blockstats.append(d)
    panel = pd.concat(rows, ignore_index=True)
    panel.to_csv(os.path.join(BUILD, "block_panel_pilot.csv"), index=False)
    pd.DataFrame(blockstats).to_csv(
        os.path.join(BUILD, "block_diagnostics.csv"), index=False)
    return panel


ANALYZE = {
    "2011-06": "june2011_analyze_table.csv",
    "2011-12": "All-NBM-Analyze-Table-December-2011/Dec2011_analyze_table.xls",
    "2012-06": "All-NBM-Analyze-Table-June-2012/Jun2012_analyze_master.xls",
    "2012-12": "All-NBM-Analyze-Table-December-2012/Dec2012_analyze_master.xlsx",
    "2013-06": "All-NBM-Analyze-Table-June-2013/Analyze_Table_Jun2013.xlsx",
    "2013-12": "All-NBM-Analyze-Table-December-2013/Analyze_Table_Dec2013.xlsx",
    "2014-06": "All-NBM-Analyze-Table-June-2014/Analyze_Table_Jun2014.xlsx",
}

KEEP = ["geography_type", "geography_id", "geography_desc", "state_fips",
        "metric_type", "land_area", "population", "terrmobilewless",
        "anywless_nosat", "wireless_advdl_gr3000k", "wireless_advdl_gr6000k",
        "wireless_advdl_gr10000k", "wireless_advdl_gr25000k",
        "terrmobilewless_error"]


def read_analyze(path):
    p = os.path.join(DATA, path)
    if p.endswith(".csv"):
        df = pd.read_csv(p, dtype={"geography_id": str}, low_memory=False)
    elif p.endswith(".xls"):
        df = pd.read_excel(p, sheet_name=0, dtype={"geography_id": str})
    else:
        import openpyxl
        wb = openpyxl.load_workbook(p, read_only=True)
        sheet = [s for s in wb.sheetnames if "POP" in s.upper()][0]
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        hdr = [str(h) for h in next(it)]
        df = pd.DataFrame(it, columns=hdr)
        df["geography_id"] = df["geography_id"].astype(str)
    df.columns = [c.strip().lower() for c in df.columns]
    cols = [c for c in KEEP if c in df.columns]
    df = df[cols]
    if "metric_type" in df.columns:
        df = df[df["metric_type"].astype(str).str.upper().str.startswith("POP")]
    return df


def build_analyze_panel():
    cache = os.path.join(BUILD, "analyze_allgeo.csv.gz")
    if os.path.exists(cache):
        return pd.read_csv(cache, dtype={"geography_id": str}, low_memory=False)
    frames = []
    for wave, path in ANALYZE.items():
        print(f"  analyze table {wave}", flush=True)
        df = read_analyze(path)
        df.insert(0, "wave", wave)
        frames.append(df)
    allgeo = pd.concat(frames, ignore_index=True)
    for c in allgeo.columns:
        if c.startswith(("terr", "anyw", "wireless", "land", "popul")):
            allgeo[c] = pd.to_numeric(allgeo[c], errors="coerce")
    allgeo.to_csv(os.path.join(BUILD, "analyze_allgeo.csv.gz"), index=False)
    return allgeo


def main_pilot():
    print("== analyze tables ==")
    allgeo = build_analyze_panel()
    cty = allgeo[allgeo["geography_type"] == "COUNTY"].copy()
    cty["county_fips"] = cty["geography_id"].str.replace(r"\.0$", "", regex=True).str.zfill(5)

    print("== block-level pilot build ==")
    panel = build_block_panel()

    a = cty[["wave", "county_fips", "population", "terrmobilewless",
             "wireless_advdl_gr6000k", "wireless_advdl_gr10000k",
             "wireless_advdl_gr25000k"]].copy()
    a["source"] = "analyze_table"
    a = a.rename(columns={
        "population": "county_pop",
        "terrmobilewless": "share_mobile_any",
        "wireless_advdl_gr6000k": "share_anywless_t6",
        "wireless_advdl_gr10000k": "share_anywless_t7",
        "wireless_advdl_gr25000k": "share_anywless_t8"})

    b = panel.rename(columns={
        "county": "county_fips",
        "mobile_any_wt": "share_mobile_any",
        "mobile_t6_wt": "share_mobile_t6",
        "mobile_t7_wt": "share_mobile_t7",
        "mobile_t7_vzw_wt": "share_mobile_t7_vzw",
        "anywless_t7_wt": "share_anywless_t7"})
    b["source"] = "block_csv"
    bcols = ["wave", "county_fips", "county_pop", "source", "share_mobile_any",
             "share_mobile_t6", "share_mobile_t7", "share_mobile_t7_vzw",
             "share_anywless_t7", "mobile_t7_any", "mobile_any_any"]
    out = pd.concat([a, b[[c for c in bcols if c in b.columns]]],
                    ignore_index=True, sort=False)
    out = out.sort_values(["source", "wave", "county_fips"])
    out.to_csv(os.path.join(OUT, "exposure_county_panel.csv"), index=False)
    print("wrote", os.path.join(OUT, "exposure_county_panel.csv"), len(out), "rows")


# ----------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--wave", help="process a single wave (e.g. 2012-06)")
    ap.add_argument("--jobs", type=int, default=3)
    ap.add_argument("--keep-zips", action="store_true",
                    help="do not delete state-wave zips after caching")
    ap.add_argument("--assemble-only", action="store_true")
    ap.add_argument("--no-assemble", action="store_true")
    ap.add_argument("--pilot", action="store_true",
                    help="legacy pilot/analyze-table build (03_validate inputs)")
    a = ap.parse_args()

    if a.pilot:
        main_pilot()
        return
    if a.assemble_only:
        assemble()
        return
    waves = [a.wave] if a.wave else list(WAVES9)
    for w in waves:
        print(f"== wave {w} ==", flush=True)
        build_wave(w, jobs=a.jobs, keep_zips=a.keep_zips)
    if not (a.wave or a.no_assemble):
        assemble()


if __name__ == "__main__":
    main()
